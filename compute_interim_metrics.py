"""
Compute aggregate metrics from eval_checkpoint.json (25 completed queries)
and save interim reports.
"""

import json, os
from collections import defaultdict

# ── Load checkpoint ────────────────────────────────────────────────────────────
with open("reports/eval_checkpoint.json") as f:
    ckpt = json.load(f)

records = ckpt["records"]
N = len(records)
print(f"\n{'='*60}")
print(f"  Interim Evaluation Report  —  {N} / 100 queries completed")
print(f"  Checkpoint timestamp: {ckpt['timestamp']}")
print(f"{'='*60}\n")

# ── Helper: mean over list ──────────────────────────────────────────────────────
def mean(vals):
    return sum(vals) / len(vals) if vals else 0.0

# ── Collect flat lists ──────────────────────────────────────────────────────────
h_r5   = [r["hybrid_recall_5"]    for r in records]
h_mrr  = [r["hybrid_mrr"]         for r in records]
rr_r5  = [r["reranked_recall_5"]  for r in records]
rr_r1  = [r["reranked_recall_1"]  for r in records]
rr_r3  = [r["reranked_recall_3"]  for r in records]
rr_mrr = [r["reranked_mrr"]       for r in records]
cit_ac = [r["citation_accuracy"]  for r in records]
faith  = [r["faithfulness"]       for r in records]

# "full accuracy" = citation correct AND page correct
# citation_accuracy already captures doc_id match; check if generated_citations
# match both doc_id AND page_number against expected
def full_cit_match(r):
    if not r["generated_citations"]:
        return 0.0
    exp_doc  = r["expected_doc_id"]
    exp_page = r["expected_page_no"]
    for c in r["generated_citations"]:
        if c["doc_id"] == exp_doc and c["page_number"] == exp_page:
            return 1.0
    return 0.0

full_ac = [full_cit_match(r) for r in records]
refused = sum(1 for r in records if r.get("is_refused"))

# ── Overall metrics ─────────────────────────────────────────────────────────────
print("OVERALL METRICS")
print(f"  {'Metric':<35} {'Value':>8}")
print(f"  {'-'*45}")
print(f"  {'Hybrid   Recall@5':<35} {mean(h_r5):>8.4f}")
print(f"  {'Hybrid   MRR':<35} {mean(h_mrr):>8.4f}")
print(f"  {'Reranked Recall@1':<35} {mean(rr_r1):>8.4f}")
print(f"  {'Reranked Recall@3':<35} {mean(rr_r3):>8.4f}")
print(f"  {'Reranked Recall@5':<35} {mean(rr_r5):>8.4f}")
print(f"  {'Reranked MRR':<35} {mean(rr_mrr):>8.4f}")
print(f"  {'Citation Accuracy (doc match)':<35} {mean(cit_ac):>8.4f}")
print(f"  {'Citation Full Accuracy (doc+page)':<35} {mean(full_ac):>8.4f}")
print(f"  {'Faithfulness (LLM judge)':<35} {mean(faith):>8.4f}")
print(f"  {'Refusal Rate':<35} {refused/N:>8.4f}  ({refused}/{N})")

# ── By doc_type ─────────────────────────────────────────────────────────────────
print("\n\nBREAKDOWN BY DOC_TYPE")
print(f"  {'Type':<10} {'N':>4}  {'R@5':>6}  {'MRR':>6}  {'CitAcc':>7}  {'CitFull':>8}  {'Faith':>7}")
print(f"  {'-'*65}")
groups_type = defaultdict(list)
for r in records:
    groups_type[r["doc_type"]].append(r)

for dt in ["judgment", "act", "pov", "tax"]:
    grp = groups_type.get(dt, [])
    if not grp:
        continue
    n = len(grp)
    r5  = mean([x["reranked_recall_5"] for x in grp])
    mrr = mean([x["reranked_mrr"]      for x in grp])
    ca  = mean([x["citation_accuracy"] for x in grp])
    cf  = mean([full_cit_match(x)      for x in grp])
    fa  = mean([x["faithfulness"]      for x in grp])
    print(f"  {dt:<10} {n:>4}  {r5:>6.4f}  {mrr:>6.4f}  {ca:>7.4f}  {cf:>8.4f}  {fa:>7.4f}")

# ── By difficulty ───────────────────────────────────────────────────────────────
print("\n\nBREAKDOWN BY DIFFICULTY")
print(f"  {'Difficulty':<14} {'N':>4}  {'R@5':>6}  {'MRR':>6}  {'CitAcc':>7}  {'CitFull':>8}  {'Faith':>7}")
print(f"  {'-'*70}")
groups_diff = defaultdict(list)
for r in records:
    groups_diff[r["difficulty"]].append(r)

for diff in ["factual", "interpretive", "multi_hop", "unanswerable"]:
    grp = groups_diff.get(diff, [])
    if not grp:
        continue
    n = len(grp)
    r5  = mean([x["reranked_recall_5"] for x in grp])
    mrr = mean([x["reranked_mrr"]      for x in grp])
    ca  = mean([x["citation_accuracy"] for x in grp])
    cf  = mean([full_cit_match(x)      for x in grp])
    fa  = mean([x["faithfulness"]      for x in grp])
    print(f"  {diff:<14} {n:>4}  {r5:>6.4f}  {mrr:>6.4f}  {ca:>7.4f}  {cf:>8.4f}  {fa:>7.4f}")

# ── Per-query quick view ─────────────────────────────────────────────────────────
print("\n\nPER-QUERY SUMMARY")
print(f"  {'QID':<6} {'Type':<10} {'Diff':<14} {'R@5':>5}  {'MRR':>6}  {'Cit':>5}  {'Faith':>6}  {'Status'}")
print(f"  {'-'*72}")
for r in records:
    ref = "REFUSED" if r.get("is_refused") else "ok"
    print(f"  {r['query_id']:<6} {r['doc_type']:<10} {r['difficulty']:<14} "
          f"{r['reranked_recall_5']:>5.2f}  {r['reranked_mrr']:>6.4f}  "
          f"{r['citation_accuracy']:>5.2f}  {r['faithfulness']:>6.4f}  {ref}")

# ── Build interim JSON report ────────────────────────────────────────────────────
def group_stats(grp):
    if not grp:
        return {}
    return {
        "n": len(grp),
        "reranked_recall_5": round(mean([x["reranked_recall_5"] for x in grp]), 4),
        "reranked_mrr":      round(mean([x["reranked_mrr"]      for x in grp]), 4),
        "citation_accuracy": round(mean([x["citation_accuracy"] for x in grp]), 4),
        "citation_full_accuracy": round(mean([full_cit_match(x) for x in grp]), 4),
        "faithfulness":      round(mean([x["faithfulness"]      for x in grp]), 4),
    }

report = {
    "meta": {
        "checkpoint_timestamp": ckpt["timestamp"],
        "completed_queries": N,
        "total_queries": 100,
        "completion_pct": N / 100 * 100,
        "refused_count": refused,
    },
    "overall": {
        "hybrid_recall_5":         round(mean(h_r5),   4),
        "hybrid_mrr":              round(mean(h_mrr),  4),
        "reranked_recall_1":       round(mean(rr_r1),  4),
        "reranked_recall_3":       round(mean(rr_r3),  4),
        "reranked_recall_5":       round(mean(rr_r5),  4),
        "reranked_mrr":            round(mean(rr_mrr), 4),
        "citation_accuracy":       round(mean(cit_ac), 4),
        "citation_full_accuracy":  round(mean(full_ac),4),
        "faithfulness":            round(mean(faith),  4),
        "refusal_rate":            round(refused / N,  4),
    },
    "by_doc_type": {
        dt: group_stats(grp) for dt, grp in groups_type.items()
    },
    "by_difficulty": {
        diff: group_stats(grp) for diff, grp in groups_diff.items()
    },
    "records": records,
}

os.makedirs("reports", exist_ok=True)
with open("reports/interim_evaluation_25.json", "w") as f:
    json.dump(report, f, indent=2)
print(f"\n  [SAVED] reports/interim_evaluation_25.json")

# ── Excel report ─────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Overall ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overall Metrics"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    sub_fill    = PatternFill("solid", fgColor="D6E4F7")

    ws1.append(["Interim Evaluation Report — 25 / 100 Queries"])
    ws1["A1"].font = Font(bold=True, size=13)
    ws1.append([f"Checkpoint: {ckpt['timestamp']}"])
    ws1.append([])

    headers = ["Metric", "Value"]
    ws1.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("Hybrid Recall@5",                  mean(h_r5)),
        ("Hybrid MRR",                        mean(h_mrr)),
        ("Reranked Recall@1",                 mean(rr_r1)),
        ("Reranked Recall@3",                 mean(rr_r3)),
        ("Reranked Recall@5",                 mean(rr_r5)),
        ("Reranked MRR",                      mean(rr_mrr)),
        ("Citation Accuracy (doc match %)",   mean(cit_ac)),
        ("Citation Full Accuracy (doc+page)", mean(full_ac)),
        ("Faithfulness (LLM judge)",          mean(faith)),
        ("Refusal Rate",                      refused / N),
    ]
    for i, (metric, val) in enumerate(rows, 5):
        ws1.cell(row=i, column=1, value=metric)
        ws1.cell(row=i, column=2, value=round(val, 4))
        ws1.cell(row=i, column=2).number_format = "0.0000"
        if i % 2 == 0:
            for c in [1, 2]:
                ws1.cell(row=i, column=c).fill = sub_fill

    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 12

    # ── Sheet 2: By Doc Type ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Doc Type")
    cols = ["Doc Type", "N", "Recall@5", "MRR", "Cit Acc", "Cit Full Acc", "Faithfulness"]
    ws2.append(cols)
    for c, h in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, (dt, grp) in enumerate(groups_type.items(), 2):
        s = group_stats(grp)
        ws2.append([dt, s["n"], s["reranked_recall_5"], s["reranked_mrr"],
                    s["citation_accuracy"], s["citation_full_accuracy"], s["faithfulness"]])
        if i % 2 == 0:
            for c in range(1, 8):
                ws2.cell(row=i, column=c).fill = sub_fill
    for c in range(1, 8):
        ws2.column_dimensions[get_column_letter(c)].width = 16

    # ── Sheet 3: By Difficulty ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Difficulty")
    ws3.append(cols[:1] + cols[1:])
    ws3["A1"].value = "Difficulty"
    for c, h in enumerate(cols, 1):
        cell = ws3.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, (diff, grp) in enumerate(groups_diff.items(), 2):
        s = group_stats(grp)
        ws3.append([diff, s["n"], s["reranked_recall_5"], s["reranked_mrr"],
                    s["citation_accuracy"], s["citation_full_accuracy"], s["faithfulness"]])
        if i % 2 == 0:
            for c in range(1, 8):
                ws3.cell(row=i, column=c).fill = sub_fill
    for c in range(1, 8):
        ws3.column_dimensions[get_column_letter(c)].width = 16

    # ── Sheet 4: Per Query ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Per Query")
    q_cols = ["Query ID", "Doc Type", "Difficulty", "Answerable",
              "Hybrid R@5", "Hybrid MRR", "Reranked R@5", "Reranked MRR",
              "Citation Acc", "Cit Full Acc", "Faithfulness", "Refused", "Status"]
    ws4.append(q_cols)
    for c, h in enumerate(q_cols, 1):
        cell = ws4.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(records, 2):
        ws4.append([
            r["query_id"], r["doc_type"], r["difficulty"], r["is_answerable"],
            r["hybrid_recall_5"], r["hybrid_mrr"],
            r["reranked_recall_5"], r["reranked_mrr"],
            r["citation_accuracy"], full_cit_match(r),
            r["faithfulness"], r.get("is_refused", False), r["status"]
        ])
        if i % 2 == 0:
            for c in range(1, len(q_cols)+1):
                ws4.cell(row=i, column=c).fill = sub_fill
    for c in range(1, len(q_cols)+1):
        ws4.column_dimensions[get_column_letter(c)].width = 14

    wb.save("reports/interim_evaluation_25.xlsx")
    print("  [SAVED] reports/interim_evaluation_25.xlsx")

except ImportError:
    print("  [SKIP]  openpyxl not installed — skipping Excel export")

print(f"\n{'='*60}\n  Done.\n{'='*60}\n")
