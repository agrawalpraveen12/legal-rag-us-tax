"""
P8 - RAG System Evaluation Runner
==================================
Runs the 100 golden queries against the Legal RAG pipeline,
computes metrics, handles rate-limit checkpointing, and
exports reports to JSON and formatted Excel.
"""

import os
import sys
import json
import time
import argparse
import pandas as pd
import numpy as np
from datetime import datetime

# Setup Python paths to allow importing from retrieve and generate modules
current_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.dirname(current_dir)
root_dir = os.path.dirname(src_dir)

sys.path.insert(0, root_dir)
sys.path.insert(0, src_dir)

from retrieve.hybrid import hybrid_search
from retrieve.rerank import rerank
from generate.answer import generate_answer
from eval.metrics import (
    calculate_recall_at_k,
    calculate_mrr,
    calculate_citation_accuracy,
    calculate_faithfulness_llm,
    calculate_cohens_h,
)

GOLDEN_EXCEL = "data/golden/golden_set_reviewed.xlsx"
OUTPUT_JSON = "reports/evaluation_results.json"
OUTPUT_EXCEL = "reports/evaluation_report.xlsx"
CHECKPOINT_FILE = "reports/eval_checkpoint.json"

def parse_args():
    parser = argparse.ArgumentParser(description="Evaluate Legal RAG System")
    parser.add_argument("--dry-run", action="store_true", help="Run evaluation on only first 2 queries")
    parser.add_argument("--no-cache", action="store_true", help="Start evaluation fresh, ignore checkpoints")
    parser.add_argument("--checkpoint-interval", type=int, default=5, help="Number of rows between checkpoints")
    return parser.parse_args()


def run_evaluation():
    args = parse_args()
    os.makedirs("reports", exist_ok=True)
    
    print("=" * 60)
    print("STARTING LEGAL RAG EVALUATION")
    print("=" * 60)
    
    # 1. Load Golden Set
    if not os.path.exists(GOLDEN_EXCEL):
        print(f"Error: Golden set file not found at {GOLDEN_EXCEL}")
        sys.exit(1)
        
    df = pd.read_excel(GOLDEN_EXCEL)
    total_queries = len(df)
    print(f"Loaded {total_queries} queries from golden set.")
    
    if args.dry_run:
        df = df.head(2)
        print(f"DRY RUN MODE: Evaluating only the first {len(df)} queries.")
        
    # 2. Checkpoint Loading
    processed_records = []
    processed_ids = set()
    
    if not args.no_cache and os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
                checkpoint_data = json.load(f)
                # Check if it was a dry-run checkpoint or matches current setup
                processed_records = checkpoint_data.get("records", [])
                processed_ids = {r["query_id"] for r in processed_records}
                print(f"Resuming from checkpoint. Loaded {len(processed_records)} completed queries.")
        except Exception as e:
            print(f"Failed to load checkpoint ({e}), starting fresh.")
            processed_records = []
            processed_ids = set()
            
    # 3. Main Evaluation Loop
    start_time = time.time()
    
    for idx, row in df.iterrows():
        query_id = row["query_id"]
        
        # Skip if already processed
        if query_id in processed_ids:
            continue
            
        query = row["query"]
        doc_type = row["doc_type"]
        difficulty = row["difficulty"]
        ground_truth = row["ground_truth_answer"]
        expected_citation_str = row["expected_citations"]
        is_answerable_str = str(row["is_answerable"]).strip().lower()
        is_answerable = (is_answerable_str == "yes")
        
        expected_doc_id = str(row["source_doc"]).strip()
        expected_page_no = int(row["page_no"])
        
        print(f"\n[{idx+1}/{len(df)}] Processing Query ID: {query_id} ({difficulty} | {doc_type})")
        print(f"Query: {query[:80]}...")
        
        try:
            # Step A: Retrieval (Hybrid Search)
            print("  Retrieving chunks...")
            hybrid_res = hybrid_search(query, rewrite=True)
            rewritten_query = hybrid_res.get("rewritten_query", query)
            hybrid_chunks = hybrid_res.get("results", [])
            
            # Step B: Reranking
            print(f"  Reranking {len(hybrid_chunks)} candidates...")
            reranked_chunks = rerank(query, hybrid_chunks, top_k=8)
            
            # Step C: Generation (all 8 reranked chunks — 70B handles the full context)
            print("  Generating answer...")
            gen_res = generate_answer(query, reranked_chunks)
            generated_answer = gen_res.get("answer", "")
            generated_citations = gen_res.get("citations", [])
            is_refused = gen_res.get("is_refused", False)
            
            # Calculate retrieval metrics (pre-rerank)
            hybrid_recall_1 = calculate_recall_at_k(hybrid_chunks, expected_doc_id, expected_page_no, 1)
            hybrid_recall_3 = calculate_recall_at_k(hybrid_chunks, expected_doc_id, expected_page_no, 3)
            hybrid_recall_5 = calculate_recall_at_k(hybrid_chunks, expected_doc_id, expected_page_no, 5)
            hybrid_recall_10 = calculate_recall_at_k(hybrid_chunks, expected_doc_id, expected_page_no, 10)
            hybrid_recall_50 = calculate_recall_at_k(hybrid_chunks, expected_doc_id, expected_page_no, 50)
            hybrid_mrr_score = calculate_mrr(hybrid_chunks, expected_doc_id, expected_page_no)
            
            # Calculate retrieval metrics (post-rerank)
            reranked_recall_1 = calculate_recall_at_k(reranked_chunks, expected_doc_id, expected_page_no, 1)
            reranked_recall_3 = calculate_recall_at_k(reranked_chunks, expected_doc_id, expected_page_no, 3)
            reranked_recall_5 = calculate_recall_at_k(reranked_chunks, expected_doc_id, expected_page_no, 5)
            reranked_recall_8 = calculate_recall_at_k(reranked_chunks, expected_doc_id, expected_page_no, 8)
            reranked_mrr_score = calculate_mrr(reranked_chunks, expected_doc_id, expected_page_no)
            
            # Calculate generation metrics
            citation_accuracy = calculate_citation_accuracy(
                generated_citations, expected_doc_id, expected_page_no, is_answerable, is_refused
            )
            
            # LLM-as-judge faithfulness (1 Groq call vs DeBERTa's strict NLI scoring)
            faithfulness = calculate_faithfulness_llm(generated_answer, reranked_chunks)
            
            record = {
                "query_id": query_id,
                "query": query,
                "rewritten_query": rewritten_query,
                "doc_type": doc_type,
                "difficulty": difficulty,
                "is_answerable": is_answerable_str,
                "expected_citation": expected_citation_str,
                "expected_doc_id": expected_doc_id,
                "expected_page_no": expected_page_no,
                "generated_answer": generated_answer,
                "generated_citations": generated_citations,
                "is_refused": is_refused,
                "hybrid_recall_1": hybrid_recall_1,
                "hybrid_recall_3": hybrid_recall_3,
                "hybrid_recall_5": hybrid_recall_5,
                "hybrid_recall_10": hybrid_recall_10,
                "hybrid_recall_50": hybrid_recall_50,
                "hybrid_mrr": hybrid_mrr_score,
                "reranked_recall_1": reranked_recall_1,
                "reranked_recall_3": reranked_recall_3,
                "reranked_recall_5": reranked_recall_5,
                "reranked_recall_8": reranked_recall_8,
                "reranked_mrr": reranked_mrr_score,
                "citation_accuracy": citation_accuracy,
                "faithfulness": faithfulness,
                "status": "success",
                "error": None
            }
            
            print(f"  -> Citation Acc: {citation_accuracy:.1%}, Faithfulness: {faithfulness:.1%}")
            
        except Exception as e:
            print(f"  -> Error processing query: {e}")
            record = {
                "query_id": query_id,
                "query": query,
                "rewritten_query": "",
                "doc_type": doc_type,
                "difficulty": difficulty,
                "is_answerable": is_answerable_str,
                "expected_citation": expected_citation_str,
                "expected_doc_id": expected_doc_id,
                "expected_page_no": expected_page_no,
                "generated_answer": f"ERROR: {str(e)}",
                "generated_citations": [],
                "is_refused": False,
                "hybrid_recall_1": 0.0,
                "hybrid_recall_3": 0.0,
                "hybrid_recall_5": 0.0,
                "hybrid_recall_10": 0.0,
                "hybrid_recall_50": 0.0,
                "hybrid_mrr": 0.0,
                "reranked_recall_1": 0.0,
                "reranked_recall_3": 0.0,
                "reranked_recall_5": 0.0,
                "reranked_recall_8": 0.0,
                "reranked_mrr": 0.0,
                "citation_accuracy": 0.0,
                "faithfulness": 0.0,
                "status": "failed",
                "error": str(e)
            }
            
        processed_records.append(record)
        processed_ids.add(query_id)
        
        # Save Checkpoint
        if len(processed_records) % args.checkpoint_interval == 0 or len(processed_records) == len(df):
            print(f"\nSaving checkpoint ({len(processed_records)} queries processed)...")
            checkpoint_state = {
                "timestamp": datetime.now().isoformat(),
                "records": processed_records,
                "dry_run": args.dry_run
            }
            with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
                json.dump(checkpoint_state, f, indent=2, ensure_ascii=False)
                
        # API cooldown: 45s covers increased token usage from 70B generation (8 chunks)
        # + 2 variant queries + LLM faithfulness call per query.
        time.sleep(45)
        
    # 4. Aggregate Performance Calculations
    print("\n" + "=" * 60)
    print("PROCESSING COMPLETED. GENERATING AGGREGATED METRICS")
    print("=" * 60)
    
    eval_df = pd.DataFrame(processed_records)
    
    # Exclude failed records from metric averages
    valid_df = eval_df[eval_df["status"] == "success"]
    if valid_df.empty:
        print("Error: All query processing failed. Cannot compute metrics.")
        sys.exit(1)
        
    overall_metrics = compute_averages(valid_df)
    overall_metrics["count"] = len(valid_df)

    # Slice by difficulty (explicit loop - avoids pandas version quirks)
    by_difficulty = {}
    for diff_val, grp in valid_df.groupby("difficulty"):
        d_metrics = compute_averages(grp)
        d_metrics["count"] = int(len(grp))
        by_difficulty[diff_val] = d_metrics

    # Slice by doc_type
    by_doc_type = {}
    for doc_val, grp in valid_df.groupby("doc_type"):
        dt_metrics = compute_averages(grp)
        dt_metrics["count"] = int(len(grp))
        by_doc_type[doc_val] = dt_metrics
    
    # 5. Cohen's h Comparisons
    cohens_h_results = calculate_cohens_h_matrix(valid_df)
    
    # 6. Save JSON Results
    results_payload = {
        "timestamp": datetime.now().isoformat(),
        "overall_metrics": overall_metrics,
        "by_difficulty": by_difficulty,
        "by_doc_type": by_doc_type,
        "cohens_h": cohens_h_results,
        "detailed_queries": processed_records
    }
    
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results_payload, f, indent=2, ensure_ascii=False)
    print(f"Saved evaluation JSON results to: {OUTPUT_JSON}")
    
    # 7. Export Excel Report
    export_excel_report(overall_metrics, by_difficulty, by_doc_type, cohens_h_results, processed_records)
    print(f"Saved evaluation Excel report to: {OUTPUT_EXCEL}")
    
    # 8. Clean up checkpoint
    if os.path.exists(CHECKPOINT_FILE) and len(valid_df) == len(df):
        try:
            os.remove(CHECKPOINT_FILE)
            print("Checkpoint cleaned up.")
        except Exception:
            pass
            
    # Print summary to console
    print("\n" + "=" * 60)
    print("EVALUATION RUN COMPLETE SUMMARY")
    print("=" * 60)
    print(f"Total Time:      {(time.time() - start_time)/60:.2f} mins")
    print(f"Queries:         {overall_metrics['count']}")
    print(f"Recall@8:        {overall_metrics['reranked_recall_8']:.1%}")
    print(f"MRR (Reranked):  {overall_metrics['reranked_mrr']:.3f}")
    print(f"Citation Acc:    {overall_metrics['citation_accuracy']:.1%}")
    print(f"Faithfulness:    {overall_metrics['faithfulness']:.1%}")
    print(f"Refusal Rate:    {overall_metrics['refusal_rate']:.1%}")
    print("=" * 60)


def compute_averages(group_df: pd.DataFrame) -> dict:
    """Helper to calculate average values of RAG metrics for a DataFrame group."""
    is_ans = group_df["is_answerable"] == "yes"
    is_unans = group_df["is_answerable"] == "no"
    
    citation_accuracy_overall = group_df["citation_accuracy"].mean()
    citation_accuracy_ans = group_df[is_ans]["citation_accuracy"].mean() if is_ans.any() else 0.0
    refusal_accuracy = group_df[is_unans]["citation_accuracy"].mean() if is_unans.any() else 0.0
    
    faithfulness_overall = group_df["faithfulness"].mean()
    faithfulness_ans = group_df[is_ans]["faithfulness"].mean() if is_ans.any() else 0.0
    
    refusal_rate = group_df["is_refused"].mean()
    correct_refusal_rate = group_df[is_unans]["is_refused"].mean() if is_unans.any() else 0.0
    incorrect_refusal_rate = group_df[is_ans]["is_refused"].mean() if is_ans.any() else 0.0
    
    return {
        "hybrid_recall_1": float(group_df["hybrid_recall_1"].mean()),
        "hybrid_recall_3": float(group_df["hybrid_recall_3"].mean()),
        "hybrid_recall_5": float(group_df["hybrid_recall_5"].mean()),
        "hybrid_recall_10": float(group_df["hybrid_recall_10"].mean()),
        "hybrid_recall_50": float(group_df["hybrid_recall_50"].mean()),
        "hybrid_mrr": float(group_df["hybrid_mrr"].mean()),
        
        "reranked_recall_1": float(group_df["reranked_recall_1"].mean()),
        "reranked_recall_3": float(group_df["reranked_recall_3"].mean()),
        "reranked_recall_5": float(group_df["reranked_recall_5"].mean()),
        "reranked_recall_8": float(group_df["reranked_recall_8"].mean()),
        "reranked_mrr": float(group_df["reranked_mrr"].mean()),
        
        "citation_accuracy": float(citation_accuracy_overall),
        "citation_accuracy_answerable": float(citation_accuracy_ans),
        "refusal_accuracy_unanswerable": float(refusal_accuracy),
        
        "faithfulness": float(faithfulness_overall),
        "faithfulness_answerable": float(faithfulness_ans),
        
        "refusal_rate": float(refusal_rate),
        "correct_refusal_rate": float(correct_refusal_rate),
        "incorrect_refusal_rate": float(incorrect_refusal_rate)
    }


def calculate_cohens_h_matrix(valid_df: pd.DataFrame) -> list:
    """Compare performance across difficulty buckets using Cohen's h."""
    comparisons = []
    
    difficulty_groups = {}
    for diff in ["factual", "interpretive", "multi_hop"]:
        difficulty_groups[diff] = valid_df[valid_df["difficulty"] == diff]
        
    pairs = [
        ("factual", "interpretive"),
        ("factual", "multi_hop"),
        ("interpretive", "multi_hop")
    ]
    
    def interpret_h(h_val):
        abs_h = abs(h_val)
        if abs_h < 0.2:
            return "Negligible (< 0.2)"
        elif abs_h < 0.5:
            return "Small Effect (0.2 - 0.5)"
        elif abs_h < 0.8:
            return "Medium Effect (0.5 - 0.8)"
        else:
            return "Large Effect (>= 0.8)"
            
    for g1, g2 in pairs:
        df1 = difficulty_groups[g1]
        df2 = difficulty_groups[g2]
        
        if df1.empty or df2.empty:
            continue
            
        # 1. Compare Citation Accuracy
        p1_cit = df1["citation_accuracy"].mean()
        p2_cit = df2["citation_accuracy"].mean()
        h_cit = calculate_cohens_h(p1_cit, p2_cit)
        
        comparisons.append({
            "comparison": f"{g1.capitalize()} vs {g2.capitalize()}",
            "metric": "Citation Accuracy",
            "p1_group": g1,
            "p1_value": float(p1_cit),
            "p2_group": g2,
            "p2_value": float(p2_cit),
            "cohens_h": float(h_cit),
            "interpretation": interpret_h(h_cit)
        })
        
        # 2. Compare Recall@8
        p1_rec = df1["reranked_recall_8"].mean()
        p2_rec = df2["reranked_recall_8"].mean()
        h_rec = calculate_cohens_h(p1_rec, p2_rec)
        
        comparisons.append({
            "comparison": f"{g1.capitalize()} vs {g2.capitalize()}",
            "metric": "Retrieval Recall@8",
            "p1_group": g1,
            "p1_value": float(p1_rec),
            "p2_group": g2,
            "p2_value": float(p2_rec),
            "cohens_h": float(h_rec),
            "interpretation": interpret_h(h_rec)
        })
        
    return comparisons


def style_worksheet(ws, title_text=None, header_row=1):
    """Applies premium formatting (Segoe UI, dark blue header, borders, number formats)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=10)
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    if title_text:
        ws.cell(row=1, column=1, value=title_text).font = title_font
        ws.row_dimensions[1].height = 30
        start_row = 3
    else:
        start_row = header_row
        
    # Format Header Row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 25
    
    # Format Data Rows
    for row in range(start_row + 1, ws.max_row + 1):
        is_zebra = (row % 2 == 0)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = thin_border
            if is_zebra:
                cell.fill = zebra_fill
            
            # Alignments & Formats based on cell values
            val = cell.value
            col_name = str(ws.cell(row=start_row, column=col).value or "")
            
            if isinstance(val, (int, float)):
                # Detect percentage metrics
                pct_headers = [
                    "recall", "accuracy", "faithfulness", "rate", 
                    "proportion", "p1", "p2", "value"
                ]
                
                # Check if it needs percent format (e.g. 0.0%)
                is_pct = any(h in col_name.lower() for h in pct_headers)
                
                # Exclude specific non-percentage fields (like ID numbers, counts, Cohen's h)
                if "count" in col_name.lower() or "id" in col_name.lower() or "page" in col_name.lower() or "cohens_h" in col_name.lower():
                    is_pct = False
                    
                if is_pct:
                    cell.number_format = '0.0%'
                elif isinstance(val, float):
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1 and title_text:
                continue
            # Handle list/dict strings beautifully in detail view
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)


def export_excel_report(overall, difficulty, doc_type, cohens_h, detailed):
    """Export aggregated evaluation results into a multi-tab formatted Excel sheet."""
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        
        # 1. Dashboard / Summary
        summary_rows = [
            {"Metric Group": "Metadata", "Metric Name": "Total Evaluated Queries", "Overall Value": overall["count"]},
            {"Metric Group": "Retrieval (Pre-Rerank)", "Metric Name": "Hybrid Recall@1", "Overall Value": overall["hybrid_recall_1"]},
            {"Metric Group": "Retrieval (Pre-Rerank)", "Metric Name": "Hybrid Recall@3", "Overall Value": overall["hybrid_recall_3"]},
            {"Metric Group": "Retrieval (Pre-Rerank)", "Metric Name": "Hybrid Recall@5", "Overall Value": overall["hybrid_recall_5"]},
            {"Metric Group": "Retrieval (Pre-Rerank)", "Metric Name": "Hybrid Recall@10", "Overall Value": overall["hybrid_recall_10"]},
            {"Metric Group": "Retrieval (Pre-Rerank)", "Metric Name": "Hybrid Recall@50", "Overall Value": overall["hybrid_recall_50"]},
            {"Metric Group": "Retrieval (Pre-Rerank)", "Metric Name": "Hybrid MRR", "Overall Value": overall["hybrid_mrr"]},
            
            {"Metric Group": "Retrieval (Post-Rerank)", "Metric Name": "Reranked Recall@1", "Overall Value": overall["reranked_recall_1"]},
            {"Metric Group": "Retrieval (Post-Rerank)", "Metric Name": "Reranked Recall@3", "Overall Value": overall["reranked_recall_3"]},
            {"Metric Group": "Retrieval (Post-Rerank)", "Metric Name": "Reranked Recall@5", "Overall Value": overall["reranked_recall_5"]},
            {"Metric Group": "Retrieval (Post-Rerank)", "Metric Name": "Reranked Recall@8", "Overall Value": overall["reranked_recall_8"]},
            {"Metric Group": "Retrieval (Post-Rerank)", "Metric Name": "Reranked MRR", "Overall Value": overall["reranked_mrr"]},
            
            {"Metric Group": "Generation", "Metric Name": "Citation Accuracy (Overall)", "Overall Value": overall["citation_accuracy"]},
            {"Metric Group": "Generation", "Metric Name": "Citation Accuracy (Answerable)", "Overall Value": overall["citation_accuracy_answerable"]},
            {"Metric Group": "Generation", "Metric Name": "Refusal Accuracy (Unanswerable)", "Overall Value": overall["refusal_accuracy_unanswerable"]},
            
            {"Metric Group": "Generation", "Metric Name": "Faithfulness (Overall)", "Overall Value": overall["faithfulness"]},
            {"Metric Group": "Generation", "Metric Name": "Faithfulness (Answerable)", "Overall Value": overall["faithfulness_answerable"]},
            
            {"Metric Group": "Refusals", "Metric Name": "Overall Refusal Rate", "Overall Value": overall["refusal_rate"]},
            {"Metric Group": "Refusals", "Metric Name": "Correct Refusal Rate (Unanswerable)", "Overall Value": overall["correct_refusal_rate"]},
            {"Metric Group": "Refusals", "Metric Name": "Incorrect Refusal Rate (Answerable)", "Overall Value": overall["incorrect_refusal_rate"]}
        ]
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2)
        
        # 2. Slice by Difficulty
        diff_rows = []
        for diff_name, metrics in difficulty.items():
            diff_rows.append({
                "Difficulty": diff_name.capitalize(),
                "Query Count": metrics["count"],
                "Reranked Recall@8": metrics["reranked_recall_8"],
                "Reranked MRR": metrics["reranked_mrr"],
                "Citation Accuracy": metrics["citation_accuracy"],
                "Faithfulness": metrics["faithfulness"],
                "Refusal Rate": metrics["refusal_rate"]
            })
        diff_df = pd.DataFrame(diff_rows)
        diff_df.to_excel(writer, sheet_name="By Difficulty", index=False, startrow=2)
        
        # 3. Slice by Doc Type
        doc_rows = []
        for doc_name, metrics in doc_type.items():
            doc_rows.append({
                "Document Type": doc_name.upper(),
                "Query Count": metrics["count"],
                "Reranked Recall@8": metrics["reranked_recall_8"],
                "Reranked MRR": metrics["reranked_mrr"],
                "Citation Accuracy": metrics["citation_accuracy"],
                "Faithfulness": metrics["faithfulness"],
                "Refusal Rate": metrics["refusal_rate"]
            })
        doc_df = pd.DataFrame(doc_rows)
        doc_df.to_excel(writer, sheet_name="By Doc Type", index=False, startrow=2)
        
        # 4. Cohen's h Matrix
        ch_df = pd.DataFrame(cohens_h)
        ch_df = ch_df.rename(columns={
            "comparison": "Comparison Pair",
            "metric": "Evaluated Metric",
            "p1_group": "Group 1",
            "p1_value": "Group 1 Prop (p1)",
            "p2_group": "Group 2",
            "p2_value": "Group 2 Prop (p2)",
            "cohens_h": "Cohen's h",
            "interpretation": "Effect Size Interpretation"
        })
        ch_df.to_excel(writer, sheet_name="Cohen's h", index=False, startrow=2)
        
        # 5. Detailed row-by-row logs
        det_rows = []
        for r in detailed:
            # Flatten citations dictionary to title, p.page format for readable Excel
            flat_cits = [f"{c['doc_title']} (p.{c['page_number']})" for c in r.get("generated_citations", [])]
            flat_cits_str = ", ".join(flat_cits) if flat_cits else "None"
            
            det_rows.append({
                "Query ID": r["query_id"],
                "Difficulty": r["difficulty"],
                "Doc Type": r["doc_type"],
                "Is Answerable": r["is_answerable"],
                "Query": r["query"],
                "Rewritten Query": r["rewritten_query"],
                "Expected Citation": r["expected_citation"],
                "LLM Generated Answer": r["generated_answer"],
                "Generated Citations": flat_cits_str,
                "Is Refused": r["is_refused"],
                "Recall@8": r["reranked_recall_8"],
                "MRR": r["reranked_mrr"],
                "Citation Accuracy": r["citation_accuracy"],
                "Faithfulness": r["faithfulness"]
            })
        det_df = pd.DataFrame(det_rows)
        det_df.to_excel(writer, sheet_name="Detailed Queries", index=False, startrow=2)
        
        # Style all tabs
        workbook = writer.book
        style_worksheet(workbook["Summary"], title_text="Legal RAG Evaluation Summary Dashboard")
        style_worksheet(workbook["By Difficulty"], title_text="Evaluation Metrics by Query Difficulty")
        style_worksheet(workbook["By Doc Type"], title_text="Evaluation Metrics by Document Type")
        style_worksheet(workbook["Cohen's h"], title_text="Proportional Analysis (Cohen's h)")
        style_worksheet(workbook["Detailed Queries"], title_text="Granular Query Evaluation Logs")


if __name__ == "__main__":
    run_evaluation()
